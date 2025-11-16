import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import pyodbc
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, Alignment 
from openpyxl.utils import get_column_letter

# Import hằng số và tiện ích
from constants import (
    APP_DARK_BLUE, APP_LIGHT_GREY, COLOR_WHITE, APP_YELLOW,
    APP_ENTRY_STYLE, APP_LABEL_STYLE, APP_BUTTON_STYLE_YELLOW,
    APP_COMBOBOX_STYLE, APP_BUTTON_STYLE_RED, APP_BUTTON_STYLE_GREEN,
    APP_COMBOBOX_STYLE
)
from gui.ui_utils import setup_themed_treeview

class GradeTab(ctk.CTkFrame):
    def __init__(self, master, db_manager):
        super().__init__(master, fg_color="transparent")
        self.db_manager = db_manager
        
        self.diem_entries = {}
        self.diem_info_labels = {}
        self.hocphan_data_diem = {} # Map { "Display String": MAHP }
        
        self._setup_layout_bangdiem()

    def _setup_layout_bangdiem(self,):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1) 

        frame_fg_color = APP_DARK_BLUE
        
        # --- 1. Frame Tìm Kiếm ---
        frame_top_diem = ctk.CTkFrame(self, fg_color=frame_fg_color, corner_radius=10)
        frame_top_diem.grid(row=0, column=0, padx=10, pady=10, sticky="new")
        frame_top_diem.columnconfigure(1, weight=1)
        
        ctk.CTkLabel(frame_top_diem, text="Tìm Mã Sinh Viên:", **APP_LABEL_STYLE).grid(row=0, column=0, padx=(15, 5), pady=10)
        entry_search_masv_diem = ctk.CTkEntry(frame_top_diem, placeholder_text="Nhập MASV (ví dụ: A123)", **APP_ENTRY_STYLE)
        entry_search_masv_diem.grid(row=0, column=1, padx=5, pady=10, sticky="ew")
        self.diem_entries['search_masv'] = entry_search_masv_diem

        ctk.CTkButton(frame_top_diem, text="Tìm SV", command=self.handle_search_student_grades, width=80, **APP_BUTTON_STYLE_YELLOW).grid(row=0, column=2, padx=10, pady=10)
        ctk.CTkButton(frame_top_diem, text="Làm Mới", command=self.handle_refresh_grades, width=80, **APP_BUTTON_STYLE_YELLOW).grid(row=0, column=3, padx=(0, 15), pady=10)
        
        # --- 2. Frame Cập Nhật & Thông Tin ---
        self.frame_edit_diem = ctk.CTkFrame(self, fg_color=frame_fg_color, corner_radius=10)
        self.frame_edit_diem.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="new")
        self.frame_edit_diem.columnconfigure((1, 3, 5), weight=1)

        # Hàng 1: Thông tin SV
        ctk.CTkLabel(self.frame_edit_diem, text="MSSV:", **APP_LABEL_STYLE).grid(row=0, column=0, padx=(15, 5), pady=8, sticky="w")
        lbl_masv = ctk.CTkLabel(self.frame_edit_diem, text="...", font=APP_LABEL_STYLE['font'], text_color=COLOR_WHITE)
        lbl_masv.grid(row=0, column=1, padx=5, pady=8, sticky="w"); self.diem_info_labels['masv'] = lbl_masv

        ctk.CTkLabel(self.frame_edit_diem, text="Họ Tên:", **APP_LABEL_STYLE).grid(row=0, column=2, padx=(15, 5), pady=8, sticky="w")
        lbl_ten = ctk.CTkLabel(self.frame_edit_diem, text="...", font=APP_LABEL_STYLE['font'], text_color=COLOR_WHITE)
        lbl_ten.grid(row=0, column=3, padx=5, pady=8, sticky="w"); self.diem_info_labels['ten'] = lbl_ten
        
        # Hàng 2: Điểm TB
        ctk.CTkLabel(self.frame_edit_diem, text="Điểm TB:", **APP_LABEL_STYLE).grid(row=1, column=0, padx=(15, 5), pady=8, sticky="w")
        lbl_dtb = ctk.CTkLabel(self.frame_edit_diem, text="...", font=APP_LABEL_STYLE['font'], text_color=APP_YELLOW)
        lbl_dtb.grid(row=1, column=1, padx=5, pady=8, sticky="w"); self.diem_info_labels['dtb'] = lbl_dtb

        ctk.CTkLabel(self.frame_edit_diem, text="Xếp Loại:", **APP_LABEL_STYLE).grid(row=1, column=2, padx=(15, 5), pady=8, sticky="w")
        lbl_xeploai = ctk.CTkLabel(self.frame_edit_diem, text="...", font=APP_LABEL_STYLE['font'], text_color=APP_YELLOW)
        lbl_xeploai.grid(row=1, column=3, padx=5, pady=8, sticky="w"); self.diem_info_labels['xeploai'] = lbl_xeploai

        # Hàng 3: Chọn Học Phần
        ctk.CTkLabel(self.frame_edit_diem, text="Chọn Học Phần:", **APP_LABEL_STYLE).grid(row=2, column=0, padx=(15, 5), pady=10, sticky="w")
        combo_mahp_diem = ctk.CTkComboBox(self.frame_edit_diem, width=400, **APP_COMBOBOX_STYLE, values=[])
        combo_mahp_diem.grid(row=2, column=1, columnspan=3, padx=5, pady=10, sticky="ew")
        self.diem_entries['mahp_combo'] = combo_mahp_diem
        self._populate_hp_combobox_diem() # Tải dữ liệu

        # Hàng 4: Nhập điểm
        ctk.CTkLabel(self.frame_edit_diem, text="Điểm Mới:", **APP_LABEL_STYLE).grid(row=3, column=0, padx=(15, 5), pady=10, sticky="w")
        entry_diem = ctk.CTkEntry(self.frame_edit_diem, width=100, placeholder_text="0-10", **APP_ENTRY_STYLE)
        entry_diem.grid(row=3, column=1, padx=5, pady=10, sticky="w"); self.diem_entries['diem'] = entry_diem

        ctk.CTkButton(self.frame_edit_diem, text="Cập Nhật Điểm", command=self.handle_update_grade, **APP_BUTTON_STYLE_YELLOW).grid(row=3, column=2, padx=15, pady=10, sticky="w")

        ctk.CTkButton(self.frame_edit_diem, text="Xóa Điểm", command=self.handle_delete_grade, **APP_BUTTON_STYLE_RED).grid(row=3, column=3, padx=15, pady=10, sticky="w")

        ctk.CTkButton(self.frame_edit_diem, text="Xuất Bảng Điểm", command=self.handle_export_grades_to_excel, **APP_BUTTON_STYLE_GREEN).grid(row=3, column=4, padx=(15, 5), pady=10, sticky="w")
        
        # --- 3. Frame Treeview  ---
        self.frame_tree_diem = ctk.CTkFrame(self, fg_color=frame_fg_color, corner_radius=10)
        self.frame_tree_diem.grid(row=2, column=0, padx=10, pady=10, sticky="nsew") 
        self._setup_diem_treeview(self.frame_tree_diem)

        # Ẩn các frame ban đầu
        self.frame_tree_diem.grid_remove()
        self.frame_edit_diem.grid_remove()

    def _setup_diem_treeview(self, parent_frame):
        columns_config = {
            "MAHP": "Mã Học Phần", "TENMH": "Tên Môn Học",
            "SOTC": "Số Tín Chỉ", "DIEM": "Điểm (Hệ 10)"
        }
        self.tree_diem = setup_themed_treeview(parent_frame, columns_config, style_name="Diem.Treeview")
        
        self.tree_diem.column("MAHP", width=120, anchor=tk.CENTER, stretch=False)
        self.tree_diem.column("TENMH", width=400, anchor=tk.W)
        self.tree_diem.column("SOTC", width=100, anchor=tk.CENTER, stretch=False)
        self.tree_diem.column("DIEM", width=120, anchor=tk.CENTER, stretch=False)
        
        self.tree_diem.bind("<<TreeviewSelect>>", self._on_diem_select)
    
    def _populate_hp_combobox_diem(self):
        self.hocphan_data_diem.clear()
        try:
            _, hocphans = self.db_manager.fetch_all_hocphan()
            # hocphans = (MAHP, MAMH, TEN_MH, HOCKY, NAMHOC, GV)
            
            display_list = []
            for hp in hocphans:
                mahp, ten_mh, hocky, namhoc = hp[0], hp[2], hp[3], hp[4]
                display_string = f"{mahp} - {ten_mh} (HK{hocky}, {namhoc})"
                
                self.hocphan_data_diem[display_string] = mahp
                display_list.append(display_string)

            self.diem_entries['mahp_combo'].configure(values=display_list)
            self.diem_entries['mahp_combo'].set("")
            
        except Exception as e:
            messagebox.showerror("Lỗi Tải Học Phần", f"Không thể tải danh sách Học Phần: {e}")
            self.diem_entries['mahp_combo'].configure(values=[])
            self.diem_entries['mahp_combo'].set("")

    def _populate_diem_treeview(self, grades_list):
        for item in self.tree_diem.get_children(): self.tree_diem.delete(item)
        for grade_row in grades_list:
            # (MAHP, TENMH, SOTC, DIEM)
            self.tree_diem.insert("", tk.END, values=grade_row)

    def _calculate_xep_loai(self, gpa):
        if gpa >= 9.0: return "Xuất sắc"
        if gpa >= 8.0: return "Giỏi"
        if gpa >= 7.0: return "Khá"
        if gpa >= 5.0: return "Trung bình"
        if gpa >= 4.0: return "Yếu"
        return "Kém"

    def _clear_diem_entries(self):
        if 'diem' in self.diem_entries:
            self.diem_entries['diem'].delete(0, 'end')
        if 'mahp_combo' in self.diem_entries:
            self.diem_entries['mahp_combo'].set("")
            
        for key, label in self.diem_info_labels.items():
            label.configure(text="...")

    def handle_search_student_grades(self):
        search_masv = self.diem_entries['search_masv'].get().strip().upper()
        if not search_masv:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập Mã Sinh Viên (MASV) để tìm.")
            return

        try:
            student_info = self.db_manager.fetch_student_info(search_masv)
            if not student_info:
                messagebox.showerror("Không tìm thấy", f"Không tìm thấy sinh viên có mã '{search_masv}'.")
                self.handle_refresh_grades()
                return
            
            student_name = student_info[0] # (TEN,)

            # Hiển thị các frame
            self.frame_tree_diem.grid()
            self.frame_edit_diem.grid()
            
            self._populate_hp_combobox_diem()
            
            _, grades = self.db_manager.fetch_grades_for_student(search_masv)
            self._populate_diem_treeview(grades)
            
            gpa_data = self.db_manager.calculate_gpa_raw(search_masv)
            gpa_str, xep_loai = "N/A", "Chưa có điểm"

            if gpa_data and gpa_data[0] not in (None, "") and gpa_data[1] not in (None, ""):
                total_weighted_score = float(gpa_data[0])
                total_credits = int(gpa_data[1])
                
                if total_credits > 0:
                    gpa = total_weighted_score / total_credits
                    gpa_str = f"{gpa:.2f}"
                    xep_loai = self._calculate_xep_loai(gpa)
                else:
                    gpa_str = "0.00"; xep_loai = "Chưa có tín chỉ"

            self._clear_diem_entries() 
            self.diem_info_labels['masv'].configure(text=search_masv)
            self.diem_info_labels['ten'].configure(text=student_name)
            self.diem_info_labels['dtb'].configure(text=gpa_str)
            self.diem_info_labels['xeploai'].configure(text=xep_loai)
            
        except Exception as e:
            messagebox.showerror("Lỗi Tìm Kiếm", f"Đã xảy ra lỗi khi tải bảng điểm: {e}")
            self.handle_refresh_grades()


    def _on_diem_select(self, event):
        selected_item = self.tree_diem.selection()
        if not selected_item: return
        
        values = self.tree_diem.item(selected_item[0], 'values')
        
        try:
            mahp_selected, diem_selected = values[0], values[3]
            
            target_display_string = None
            for display_str, mahp_id in self.hocphan_data_diem.items():
                if str(mahp_id) == str(mahp_selected):
                    target_display_string = display_str
                    break
            
            if target_display_string:
                self.diem_entries['mahp_combo'].set(target_display_string)
            else:
                self.diem_entries['mahp_combo'].set(f"LỖI: Không tìm thấy HP {mahp_selected}")

            self.diem_entries['diem'].delete(0, 'end')
            self.diem_entries['diem'].insert(0, diem_selected)

        except (IndexError, KeyError) as e:
             messagebox.showerror("Lỗi Tải Dữ Liệu", f"Không thể tải thông tin điểm. Lỗi: {e}")

    def handle_update_grade(self):
        masv = self.diem_info_labels['masv'].cget('text')
        if not masv or masv == "...":
            messagebox.showwarning("Lỗi", "Không xác định được Sinh viên. Vui lòng 'Tìm SV' trước.")
            return

        selected_display_string = self.diem_entries['mahp_combo'].get()
        if not selected_display_string:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn một Học Phần từ danh sách.")
            return
            
        try:
            mahp = self.hocphan_data_diem[selected_display_string]
        except KeyError:
            messagebox.showerror("Lỗi", f"Học phần '{selected_display_string}' không hợp lệ hoặc không tìm thấy.")
            return

        diem_str = self.diem_entries['diem'].get().strip()
        if not diem_str:
            messagebox.showwarning("Thiếu điểm", "Vui lòng nhập điểm mới (từ 0 đến 10).")
            return
            
        try:
            diem = float(diem_str)
            if not (0 <= diem <= 10):
                raise ValueError("Điểm ngoài phạm vi")
        except ValueError:
            messagebox.showwarning("Lỗi Dữ Liệu", "Điểm phải là một số hợp lệ từ 0 đến 10.")
            return
        
        try:
        # Lấy MAMH từ MAHP
            mamh = self.db_manager.get_mamh_from_mahp(mahp)
            if mamh is None:
             messagebox.showerror("Lỗi", f"Không tìm thấy Mã Môn Học cho Học Phần {mahp}.")
             return

        # Kiểm tra các môn còn thiếu (giả sử đậu >= 5.0)
        # Hàm này giờ trả về danh sách các tuple: [('dsg101', 'Nhập môn lập trình'), ...]
            missing_subjects_list = self.db_manager.check_missing_prerequisites(masv, mamh, 5.0)
        
            if missing_subjects_list:
            
            
            # 1. Lấy tên học phần mà người dùng đang cố nhập điểm
                hoc_phan_dang_chon = self.diem_entries['mahp_combo'].get()
            
            # 2. Tạo danh sách các môn thiếu (xuống dòng)
                mon_thieu_str = ""
                for (mamh_thieu, ten_mh_thieu) in missing_subjects_list:
                    mon_thieu_str += f"\n- {ten_mh_thieu} ({mamh_thieu})"
            
            # 3. Hiển thị thông báo
                messagebox.showwarning("Không đủ điều kiện", 
                                   f"Sinh viên '{masv}' không thể nhập điểm cho học phần:\n"
                                   f"'{hoc_phan_dang_chon}'\n\n"
                                   f"Lý do: Chưa hoàn thành các môn tiên quyết sau:"
                                   f"{mon_thieu_str}")
                return # Dừng, không cho cập nhật điểm

        except Exception as e:
            messagebox.showerror("Lỗi Kiểm Tra Điều Kiện", f"Không thể xác minh môn tiên quyết:\n{e}")
            return # Dừng

        try:
            self.db_manager.add_or_update_grade(masv, mahp, diem)
            messagebox.showinfo("Thành công", f"Đã cập nhật/thêm điểm cho HP {mahp} của SV {masv} thành {diem}.")
            self.handle_search_student_grades() # Tải lại
        except pyodbc.IntegrityError as e:
             error_message = str(e)
             if "FOREIGN KEY" in error_message:
                 messagebox.showwarning("Lỗi Dữ Liệu", f"Lỗi khóa ngoại: MASV '{masv}' hoặc MAHP '{mahp}' không tồn tại.")
             else:
                 messagebox.showwarning("Lỗi Toàn Vẹn", f"Lỗi khi cập nhật điểm: {e}")
        except Exception as e:
            messagebox.showerror("Lỗi Cập Nhật", f"Lỗi không xác định khi cập nhật điểm: {e}")

    def handle_delete_grade(self):
        """
        Xóa một dòng điểm (KETQUA) dựa trên SV và dòng được CHỌN trong Treeview.
        """
        
        # 1. Lấy MASV (từ label)
        masv = self.diem_info_labels['masv'].cget('text')
        if not masv or masv == "...":
            messagebox.showwarning("Lỗi", "Không xác định được Sinh viên. Vui lòng 'Tìm SV' trước.")
            return

        # 2. Lấy MAHP và Tên Môn Học từ Treeview
        selected_item = self.tree_diem.selection()
        if not selected_item:
            messagebox.showwarning("Chưa chọn", "Vui lòng click chọn một dòng điểm trong bảng để xóa.")
            return
        
        try:
            values = self.tree_diem.item(selected_item[0], 'values')
            mahp = values[0]     # MAHP là cột 0
            ten_mh = values[1]   # TENMH là cột 1
        except (IndexError, KeyError):
            messagebox.showerror("Lỗi", "Không thể lấy thông tin Mã Học Phần từ dòng đã chọn.")
            return
            
        # 3. Xác nhận xóa
        if not messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa điểm môn:\n\n{ten_mh} (HP: {mahp})\n\nCủa sinh viên {masv}?"):
            return

        # 4. Thực hiện xóa
        try:
            self.db_manager.delete_grade(masv, mahp)
            messagebox.showinfo("Thành công", f"Đã xóa điểm của HP {mahp} cho SV {masv}.")
            # Tự động tìm kiếm lại để làm mới Treeview và GPA
            self.handle_search_student_grades()
            
        except ValueError as e: # Lỗi từ db_manager (e.g., "Không tìm thấy...")
            messagebox.showwarning("Lỗi Xóa", str(e))
        except pyodbc.Error as e: # Lỗi SQL (ví dụ: quyền)
            messagebox.showerror("Lỗi Database", f"Lỗi SQL khi xóa điểm: {e}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi không xác định khi xóa điểm: {e}")


    def handle_refresh_grades(self):
        self.diem_entries['search_masv'].delete(0, 'end')
        self._clear_diem_entries()
        
        for item in self.tree_diem.get_children():
            self.tree_diem.delete(item)
                
        self.frame_tree_diem.grid_remove()
        self.frame_edit_diem.grid_remove()
            
        self._populate_hp_combobox_diem()
        
    def handle_export_grades_to_excel(self):
        """
        Xuất bảng điểm của sinh viên đang xem ra file Excel.
        """
        
        # 1. Lấy thông tin sinh viên
        masv = self.diem_info_labels['masv'].cget('text')
        ten_sv = self.diem_info_labels['ten'].cget('text')
        dtb = self.diem_info_labels['dtb'].cget('text')
        xep_loai = self.diem_info_labels['xeploai'].cget('text')
        
        # 2. Đề xuất tên file mặc định
        default_filename = f"BangDiem_{masv}_{ten_sv.replace(' ', '_')}.xlsx"

        # 3. Hỏi người dùng muốn lưu file ở đâu
        file_path = filedialog.asksaveasfilename(
            title=f"Lưu bảng điểm cho SV {masv}",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return

        try:
            # 4. Tạo Workbook và Sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Bảng Điểm {masv}"
            
            # --- Định dạng file Excel ---
            
            # Tiêu đề chính
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = "BẢNG ĐIỂM SINH VIÊN"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Thông tin sinh viên
            ws['A3'] = "Mã số sinh viên:"
            ws['B3'] = masv
            ws['A4'] = "Họ và tên:"
            ws['B4'] = ten_sv
            
            ws['C3'] = "Điểm TB (Hệ 10):"
            ws['D3'] = dtb
            ws['C4'] = "Xếp loại:"
            ws['D4'] = xep_loai
            
            # Làm đậm các nhãn thông tin
            for cell_id in ['A3', 'A4', 'C3', 'C4']:
                ws[cell_id].font = Font(bold=True)

            # 5. Tiêu đề bảng điểm
            headers = [self.tree_diem.heading(col)["text"] for col in self.tree_diem["columns"]]
            # headers = ("MAHP", "TENMH", "SOTC", "DIEM")
            ws.append([]) # Thêm 1 dòng trống
            ws.append(headers)
            
            header_row = ws[6] # Dòng tiêu đề là dòng 6
            for cell in header_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # 6. Thêm dữ liệu điểm từ Treeview
            for item_id in self.tree_diem.get_children():
                row_values = self.tree_diem.item(item_id, 'values')
                ws.append(row_values)
            
            # 7. Tự động chỉnh độ rộng cột
            column_widths = {'A': 15, 'B': 40, 'C': 15, 'D': 15} # (MAHP, TENMH, SOTC, DIEM)
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 8. Lưu file
            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất bảng điểm thành công!\nĐường dẫn: {file_path}")

        except Exception as e:
            messagebox.showerror("Lỗi Xuất Excel", f"Có lỗi xảy ra khi lưu file:\n{e}\n(Có thể file đang được mở bởi chương trình khác.)")